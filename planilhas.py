from openpyxl import load_workbook

wb=load_workbook("C:\\Users\\pedro\\Downloads\\aluos.xlsx")

ws = wb['Plan1']

for linha in range(2,5):
   aluno = ws.cell(column=1, row=linha).value
   nota1 = ws.cell(column=2, row=linha).value
   nota2 = ws.cell(column=3, row=linha).value
   print("Aluno: ", aluno)
   print("Prova-1: ", nota1)
   print("Prova-2: ", nota2)
   media = (nota1 + nota2)/2
   print("Média: ",media)
   if media >= 7:
       print("O aluno, ", aluno," foi aprovado.")
   else:
       print("O aluno, ", aluno," foi reprovado.")

print("---------------------")
